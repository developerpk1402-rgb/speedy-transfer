from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
import json
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.db.models import Sum, Count, Value, DecimalField
from django.db.models.functions import Coalesce, TruncDate
import datetime
from django.utils import timezone
from django.conf import settings
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Booking
import csv

# xlsx export
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def staff_check(user):
    return user.is_active and user.is_staff


@login_required
@user_passes_test(staff_check)
def reports_portal(request):
    return render(request, 'speedy_app/reports/portal.html')


@login_required
@user_passes_test(staff_check)
def sold_orders_report(request):
    # base queryset with related fields to avoid N+1
    qs = Booking.objects.select_related('pickup_location1', 'dropoff_location1').all().order_by('-date_capture')

    # date filters (parse dates and filter by datetime ranges to avoid timezone/date mismatches)
    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    def _normalize_dt(dt):
        """Return a datetime compatible with the project's USE_TZ setting.

        If USE_TZ is True, return an aware datetime (make_aware if necessary).
        If USE_TZ is False, return a naive datetime (strip tzinfo if necessary).
        """
        if settings.USE_TZ:
            return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        else:
            return dt if timezone.is_naive(dt) else dt.replace(tzinfo=None)

    try:
        if start:
            start_date = datetime.datetime.strptime(start, '%Y-%m-%d').date()
            start_dt = datetime.datetime.combine(start_date, datetime.time.min)
            qs = qs.filter(date_capture__gte=_normalize_dt(start_dt))
        if end:
            end_date = datetime.datetime.strptime(end, '%Y-%m-%d').date()
            # include the whole end day
            end_dt = datetime.datetime.combine(end_date, datetime.time.max)
            qs = qs.filter(date_capture__lte=_normalize_dt(end_dt))
    except Exception:
        # if parsing fails, ignore filters (alternatively we could surface an error)
        pass

    # pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(qs, 25)
    try:
        bookings = paginator.page(page)
    except PageNotAnInteger:
        bookings = paginator.page(1)
    except EmptyPage:
        bookings = paginator.page(paginator.num_pages)

    fmt = request.GET.get('format')
    # CSV export
    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sold_orders.csv"'
        writer = csv.writer(response)
        writer.writerow(['Booking ID', 'Client', 'Customer Name', 'Phone', 'Pickup', 'Dropoff', 'Pickup Date', 'Total Amount', 'Currency', 'Payment Method', 'Created'])
        for b in qs:
            writer.writerow([b.id, b.client_id, b.customer_name, b.customer_phone, getattr(b.pickup_location1, 'name', ''), getattr(b.dropoff_location1, 'name', ''), b.pickup_date_time, b.total_amount, b.currency, b.payment_method, b.date_capture])
        return response

    # XLSX export
    if fmt == 'xlsx' and OPENPYXL_AVAILABLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['Booking ID', 'Client', 'Customer Name', 'Phone', 'Pickup', 'Dropoff', 'Pickup Date', 'Total Amount', 'Currency', 'Payment Method', 'Created']
        ws.append(headers)
        for b in qs:
            ws.append([b.id, b.client_id, b.customer_name, b.customer_phone, getattr(b.pickup_location1, 'name', ''), getattr(b.dropoff_location1, 'name', ''), b.pickup_date_time.strftime('%Y-%m-%d %H:%M:%S') if b.pickup_date_time else '', float(b.total_amount) if b.total_amount else 0, b.currency, b.payment_method, b.date_capture.strftime('%Y-%m-%d %H:%M:%S')])
        # autosize
        for i, column in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sold_orders.xlsx"'
        wb.save(response)
        return response

    # return a TemplateResponse so calling code (admin wrapper) can inspect context_data
    return TemplateResponse(request, 'speedy_app/reports/sold_orders.html', {'bookings': bookings, 'paginator': paginator, 'start_date': start, 'end_date': end})


@login_required
@user_passes_test(staff_check)
def sales_history_report(request):
    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    base_qs = Booking.objects.all()
    def _normalize_dt(dt):
        if settings.USE_TZ:
            return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        else:
            return dt if timezone.is_naive(dt) else dt.replace(tzinfo=None)

    try:
        if start:
            start_date = datetime.datetime.strptime(start, '%Y-%m-%d').date()
            start_dt = datetime.datetime.combine(start_date, datetime.time.min)
            base_qs = base_qs.filter(date_capture__gte=_normalize_dt(start_dt))
        if end:
            end_date = datetime.datetime.strptime(end, '%Y-%m-%d').date()
            end_dt = datetime.datetime.combine(end_date, datetime.time.max)
            base_qs = base_qs.filter(date_capture__lte=_normalize_dt(end_dt))
    except Exception:
        pass

    # group by date using TruncDate and Coalesce to ensure totals are numeric
    qs = base_qs.annotate(date_only=TruncDate('date_capture')).values('date_only').annotate(total_sales=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField())), orders=Count('id')).order_by('-date_only')

    fmt = request.GET.get('format')
    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sales_history.csv"'
        writer = csv.writer(response)
        writer.writerow(['Date', 'Total Sales', 'Orders'])
        for row in qs:
            writer.writerow([row.get('date_only'), row.get('total_sales') or 0, row.get('orders')])
        return response

    if fmt == 'xlsx' and OPENPYXL_AVAILABLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Date', 'Total Sales', 'Orders'])
        for row in qs:
            ws.append([row.get('date_only'), float(row.get('total_sales') or 0), row.get('orders')])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sales_history.xlsx"'
        wb.save(response)
        return response

    # prepare chart data (JSON-friendly lists)
    # prepare chart data (JSON-friendly lists)
    rows = list(qs)
    chart_labels = [str(r['date_only']) for r in reversed(rows)]
    chart_values = [float(r['total_sales'] or 0) for r in reversed(rows)]

    chart_labels_json = json.dumps(chart_labels)
    chart_values_json = json.dumps(chart_values)

    # compute totals for the filtered range
    totals = base_qs.aggregate(total_sales=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField())), total_orders=Coalesce(Count('id'), Value(0)))
    total_sales = float(totals.get('total_sales') or 0)
    total_orders = int(totals.get('total_orders') or 0)
    avg_order = (total_sales / total_orders) if total_orders else 0

    # use TemplateResponse for inspectable context and deferred rendering
    return TemplateResponse(request, 'speedy_app/reports/sales_history.html', {
        'data': qs,
        'chart_labels_json': chart_labels_json,
        'chart_values_json': chart_values_json,
        'start_date': start,
        'end_date': end,
        'total_sales': total_sales,
        'total_orders': total_orders,
        'avg_order': round(avg_order,2),
    })
