from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages

from django.http import HttpResponse, HttpResponseBadRequest
from django.db.models import Sum
from django.db.models.functions import TruncDate
import csv
import datetime

from openpyxl import Workbook

try:
	from reportlab.lib.pagesizes import letter
	from reportlab.pdfgen import canvas
	REPORTLAB_AVAILABLE = True
except Exception:
	REPORTLAB_AVAILABLE = False

from speedy_app.core.models import Booking, Payment


def _is_staff(user):
	return user.is_authenticated and user.is_staff


def reports_login(request):
    """Simple login view for the reporting panel.

    Uses ReportingAuthBackend which uses specific database credentials for authentication:
    Host: 45.82.72.136:3306
    Database: speedy
    User: speedy_user
    """
    try:
        # Test database connection
        from django.db import connections
        from django.db.utils import OperationalError
        db_conn = connections['reports']
        db_conn.cursor()
        connection_successful = True
    except OperationalError:
        connection_successful = False
        messages.error(request, 'Database connection failed. Please contact support.')
    except Exception as e:
        connection_successful = False
        messages.error(request, f'Error: {str(e)}')

    if request.user.is_authenticated and request.user.is_staff:
        return redirect('reports:dashboard')

    if request.method == 'POST' and connection_successful:
        username = request.POST.get('username')
        password = request.POST.get('password')
        # Use the reporting authentication backend that connects to specific database
        user = authenticate(request, username=username, password=password,
                          backend='reports.auth.ReportingAuthBackend')
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('reports:dashboard')
        messages.error(request, 'Invalid credentials or insufficient permissions.')

    return render(request, 'reports/login.html')

@login_required(login_url='/reports/login/')
@user_passes_test(_is_staff, login_url='/reports/login/')
def reports_dashboard(request):
    """Dashboard view for the reporting panel."""
    try:
        # Test query to verify database connection
        from django.contrib.auth import get_user_model
        User = get_user_model()
        staff_count = User.objects.using('reports').filter(is_staff=True).count()
        return render(request, 'reports/dashboard.html', {'staff_count': staff_count})
    except Exception as e:
        messages.error(request, f'Database error: {str(e)}')
        return redirect('reports:login')


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def dashboard(request):
	"""Dashboard showing three tiles: Sells History, Bookings, Sold History.

	The tiles currently link to the detail report pages.
	"""
	# Provide richer tile metadata (image path, description, category, format, created)
	# Compute live metrics for each tile
	# Sells: total payments in last 30 days
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	sells_qs = Payment.objects.filter(paid_at__date__range=(start, end))
	total_sells = sells_qs.aggregate(total=Sum('amount'))['total'] or 0

	# Bookings: total bookings count
	bookings_count = Booking.objects.count()

	# Sold: combination metric (payments count + bookings count)
	payments_count = Payment.objects.count()
	sold_count = payments_count + bookings_count

	tiles = [
		{
			'title': 'Sells History',
			'url': '/reports/sells/preview/',
			'image': 'sales.svg',
			'description': 'Monthly sales performance report',
			'category': 'Sales', 
			'format': 'PDF',
			'created': '15/01/2024',
			'stat_label': 'Last 30d',
			'stat_value': f"{float(total_sells):,.2f}"
		},
		{
			'title': 'Bookings',
			'url': '/reports/bookings/preview/',
			'image': 'bookings.svg',
			'description': 'Recent bookings and details',
			'category': 'Bookings',
			'format': 'PDF',
			'created': '17/01/2024',
			'stat_label': 'Total',
			'stat_value': f"{bookings_count}"
		},
		{
			'title': 'Sold History',
			'url': '/reports/sold/preview/',
			'image': 'sold.svg',
			'description': 'Combined payments and sold items',
			'category': 'Finance',
			'format': 'PDF',
			'created': '18/01/2024',
			'stat_label': 'Items',
			'stat_value': f"{sold_count}"
		},
	]

	return render(request, 'reports/dashboard.html', {'tiles': tiles})


def reports_logout(request):
	logout(request)
	return redirect('reports:login')


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sells_history_preview(request):
	"""Preview version of sells history report."""
	response = sells_history(request)
	response['X-Frame-Options'] = 'SAMEORIGIN'  # Allow iframe embedding
	return response


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def bookings_list_preview(request):
	"""Preview version of bookings list."""
	response = bookings_list(request)
	response['X-Frame-Options'] = 'SAMEORIGIN'  # Allow iframe embedding
	return response


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sold_history_preview(request):
	"""Preview version of sold history."""
	response = sold_history(request)
	response['X-Frame-Options'] = 'SAMEORIGIN'  # Allow iframe embedding
	return response


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sells_history(request):
	"""Show aggregated payments over the last 30 days as a chart."""
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	qs = Payment.objects.filter(paid_at__date__range=(start, end))
	agg = qs.annotate(day=TruncDate('paid_at')).values('day').annotate(total=Sum('amount')).order_by('day')

	# Build mapping day -> total for full range
	totals = {item['day']: float(item['total'] or 0) for item in agg}
	labels = []
	data = []
	for i in range(30):
		d = start + datetime.timedelta(days=i)
		labels.append(d.strftime('%Y-%m-%d'))
		data.append(totals.get(d, 0))

	return render(request, 'reports/sells.html', {'labels': labels, 'data': data})


def _write_csv_response(filename, headers, rows):
	resp = HttpResponse(content_type='text/csv')
	resp['Content-Disposition'] = f'attachment; filename="{filename}"'
	writer = csv.writer(resp)
	writer.writerow(headers)
	for r in rows:
		writer.writerow(r)
	return resp


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sells_history_csv(request):
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	qs = Payment.objects.filter(paid_at__date__range=(start, end))
	agg = qs.annotate(day=TruncDate('paid_at')).values('day').annotate(total=Sum('amount')).order_by('day')
	rows = [(item['day'].strftime('%Y-%m-%d'), float(item['total'] or 0)) for item in agg]
	return _write_csv_response('sells_history.csv', ['date', 'total'], rows)


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sells_history_xlsx(request):
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	qs = Payment.objects.filter(paid_at__date__range=(start, end))
	agg = qs.annotate(day=TruncDate('paid_at')).values('day').annotate(total=Sum('amount')).order_by('day')

	wb = Workbook()
	ws = wb.active
	ws.title = 'Sells'
	ws.append(['date', 'total'])
	for item in agg:
		ws.append([item['day'].strftime('%Y-%m-%d'), float(item['total'] or 0)])

	resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	resp['Content-Disposition'] = 'attachment; filename="sells_history.xlsx"'
	wb.save(resp)
	return resp


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def bookings_list(request):
	qs = Booking.objects.all().order_by('-date_capture')
	# Simple pagination (page param)
	try:
		page = int(request.GET.get('page', '1'))
	except ValueError:
		page = 1
	per_page = 25
	start = (page - 1) * per_page
	end = start + per_page
	page_items = qs[start:end]
	return render(request, 'reports/bookings.html', {'bookings': page_items, 'page': page})


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def bookings_csv(request):
	qs = Booking.objects.all().order_by('-date_capture')
	rows = []
	for b in qs:
		rows.append([
			b.id,
			b.client_id,
			b.customer_name,
			b.customer_phone,
			b.pickup_date_time.strftime('%Y-%m-%d %H:%M'),
			b.return_date_time.strftime('%Y-%m-%d %H:%M') if b.return_date_time else '',
			float(b.total_amount or 0),
			b.currency,
		])
	return _write_csv_response('bookings.csv', ['id', 'client_id', 'customer_name', 'phone', 'pickup', 'return', 'total', 'currency'], rows)


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def bookings_xlsx(request):
	qs = Booking.objects.all().order_by('-date_capture')
	wb = Workbook()
	ws = wb.active
	ws.title = 'Bookings'
	ws.append(['id', 'client_id', 'customer_name', 'phone', 'pickup', 'return', 'total', 'currency'])
	for b in qs:
		ws.append([
			b.id,
			b.client_id,
			b.customer_name,
			b.customer_phone,
			b.pickup_date_time.strftime('%Y-%m-%d %H:%M'),
			b.return_date_time.strftime('%Y-%m-%d %H:%M') if b.return_date_time else '',
			float(b.total_amount or 0),
			b.currency,
		])
	resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	resp['Content-Disposition'] = 'attachment; filename="bookings.xlsx"'
	wb.save(resp)
	return resp


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sold_history(request):
	# Combine bookings and payments: render chart (payments) and bookings table
	# Reuse sells aggregation
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	qs = Payment.objects.filter(paid_at__date__range=(start, end))
	agg = qs.annotate(day=TruncDate('paid_at')).values('day').annotate(total=Sum('amount')).order_by('day')
	totals = {item['day']: float(item['total'] or 0) for item in agg}
	labels = []
	data = []
	for i in range(30):
		d = start + datetime.timedelta(days=i)
		labels.append(d.strftime('%Y-%m-%d'))
		data.append(totals.get(d, 0))

	bookings = Booking.objects.all().order_by('-date_capture')[:100]
	return render(request, 'reports/sold.html', {'labels': labels, 'data': data, 'bookings': bookings})


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sold_history_csv(request):
	# Combine datasets into CSV: payments first, then bookings
	# Payments
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	payments = Payment.objects.filter(paid_at__date__range=(start, end)).annotate(day=TruncDate('paid_at')).values('day').annotate(total=Sum('amount')).order_by('day')
	rows = []
	rows.append(['Payments'])
	rows.append(['date', 'total'])
	for p in payments:
		rows.append([p['day'].strftime('%Y-%m-%d'), float(p['total'] or 0)])
	rows.append([])
	rows.append(['Bookings'])
	rows.append(['id', 'client_id', 'customer_name', 'phone', 'pickup', 'return', 'total', 'currency'])
	for b in Booking.objects.all().order_by('-date_capture'):
		rows.append([
			b.id,
			b.client_id,
			b.customer_name,
			b.customer_phone,
			b.pickup_date_time.strftime('%Y-%m-%d %H:%M'),
			b.return_date_time.strftime('%Y-%m-%d %H:%M') if b.return_date_time else '',
			float(b.total_amount or 0),
			b.currency,
		])
	# Build CSV response
	resp = HttpResponse(content_type='text/csv')
	resp['Content-Disposition'] = 'attachment; filename="sold_history.csv"'
	writer = csv.writer(resp)
	for r in rows:
		writer.writerow(r)
	return resp


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sold_history_xlsx(request):
	wb = Workbook()
	ws1 = wb.active
	ws1.title = 'Payments'
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	payments = Payment.objects.filter(paid_at__date__range=(start, end)).annotate(day=TruncDate('paid_at')).values('day').annotate(total=Sum('amount')).order_by('day')
	ws1.append(['date', 'total'])
	for p in payments:
		ws1.append([p['day'].strftime('%Y-%m-%d'), float(p['total'] or 0)])

	ws2 = wb.create_sheet('Bookings')
	ws2.append(['id', 'client_id', 'customer_name', 'phone', 'pickup', 'return', 'total', 'currency'])
	for b in Booking.objects.all().order_by('-date_capture'):
		ws2.append([
			b.id,
			b.client_id,
			b.customer_name,
			b.customer_phone,
			b.pickup_date_time.strftime('%Y-%m-%d %H:%M'),
			b.return_date_time.strftime('%Y-%m-%d %H:%M') if b.return_date_time else '',
			float(b.total_amount or 0),
			b.currency,
		])

	resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	resp['Content-Disposition'] = 'attachment; filename="sold_history.xlsx"'
	wb.save(resp)
	return resp


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sells_history_pdf(request):
	if not REPORTLAB_AVAILABLE:
		return HttpResponseBadRequest('PDF export requires reportlab package. Please install reportlab.')
	# Simple PDF with list of date/total
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	qs = Payment.objects.filter(paid_at__date__range=(start, end))
	agg = qs.annotate(day=TruncDate('paid_at')).values('day').annotate(total=Sum('amount')).order_by('day')

	resp = HttpResponse(content_type='application/pdf')
	resp['Content-Disposition'] = 'attachment; filename="sells_history.pdf"'
	p = canvas.Canvas(resp, pagesize=letter)
	y = 750
	p.setFont('Helvetica-Bold', 14)
	p.drawString(40, y, 'Sells History')
	y -= 30
	p.setFont('Helvetica', 10)
	for item in agg:
		line = f"{item['day'].strftime('%Y-%m-%d')}: {float(item['total'] or 0)}"
		p.drawString(40, y, line)
		y -= 14
		if y < 40:
			p.showPage()
			y = 750
	p.showPage()
	p.save()
	return resp


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def bookings_pdf(request):
	if not REPORTLAB_AVAILABLE:
		return HttpResponseBadRequest('PDF export requires reportlab package. Please install reportlab.')
	qs = Booking.objects.all().order_by('-date_capture')
	resp = HttpResponse(content_type='application/pdf')
	resp['Content-Disposition'] = 'attachment; filename="bookings.pdf"'
	p = canvas.Canvas(resp, pagesize=letter)
	y = 750
	p.setFont('Helvetica-Bold', 14)
	p.drawString(40, y, 'Bookings')
	y -= 30
	p.setFont('Helvetica', 10)
	for b in qs:
		line = f"{b.id} | {b.client_id} | {b.customer_name} | {b.pickup_date_time.strftime('%Y-%m-%d %H:%M')} | {float(b.total_amount or 0)} {b.currency}"
		p.drawString(40, y, line[:120])
		y -= 14
		if y < 40:
			p.showPage()
			y = 750
	p.showPage()
	p.save()
	return resp


@login_required(login_url='/reports/login/')
@user_passes_test(lambda u: u.is_staff, login_url='/reports/login/')
def sold_history_pdf(request):
	if not REPORTLAB_AVAILABLE:
		return HttpResponseBadRequest('PDF export requires reportlab package. Please install reportlab.')
	# Create a PDF with payments section and bookings section
	end = datetime.date.today()
	start = end - datetime.timedelta(days=29)
	payments = Payment.objects.filter(paid_at__date__range=(start, end)).annotate(day=TruncDate('paid_at')).values('day').annotate(total=Sum('amount')).order_by('day')
	qs = Booking.objects.all().order_by('-date_capture')

	resp = HttpResponse(content_type='application/pdf')
	resp['Content-Disposition'] = 'attachment; filename="sold_history.pdf"'
	p = canvas.Canvas(resp, pagesize=letter)
	y = 750
	p.setFont('Helvetica-Bold', 14)
	p.drawString(40, y, 'Sold History - Payments')
	y -= 30
	p.setFont('Helvetica', 10)
	for pmt in payments:
		line = f"{pmt['day'].strftime('%Y-%m-%d')}: {float(pmt['total'] or 0)}"
		p.drawString(40, y, line)
		y -= 14
		if y < 80:
			p.showPage()
			y = 750
	p.showPage()
	p.save()
	return resp

